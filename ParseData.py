import openpyxl
import cosine_similarity

#wb = openpyxl.load_workbook('Code-Quality Questions.xlsx');
wb = openpyxl.load_workbook('Only Questions Java.xlsx');
print wb.sheetnames

sheet = wb['Sheet3'];
print sheet.max_row;
print sheet.max_column;

documents = list()

for row_number in range(2, sheet.max_row + 1) :
    #A -> Body (text) of the question
    body = sheet['A' + str(row_number)].value

    #B -> Title of the question
    title = sheet['B' + str(row_number)].value

    #C -> Tags of the question
    tags = sheet['C' + str(row_number)].value

    document = title + body
    documents.append(document)

documents_tuple = tuple(documents)
cosine_similarity.calculate_cosine_similarity(documents_tuple)


